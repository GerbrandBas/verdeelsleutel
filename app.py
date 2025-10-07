
import io
import math
import pandas as pd
import numpy as np
import streamlit as st

# Try matplotlib with a headless backend; if it fails, fall back to st.bar_chart
HAVE_MPL = True
MPL_ERR = ""
try:
    import matplotlib
    matplotlib.use("Agg")  # headless backend
    import matplotlib.pyplot as plt
except Exception as e:
    HAVE_MPL = False
    MPL_ERR = str(e)

from openpyxl import load_workbook

st.set_page_config(page_title="Thuiskopie Verdeelsleutel Simulator", layout="wide")

st.title("Thuiskopie Verdeelsleutel — Scenario A/B/C • met stabilisatie & 25% beeld-target")

st.markdown("""
Gebruik dit hulpmiddel om de **disciplineverdeling** te simuleren o.b.v. je Excel-model
(**Parameters**, **Baseline**, **Invoer_2023**). De app leest standaard het meegeleverde modelbestand,
maar je kunt ook een eigen Excel uploaden met dezelfde sheet- en veldnamen.
""")

# If matplotlib is unavailable, show a brief note (we will still render charts via Streamlit fallback)
if not HAVE_MPL:
    st.info("Matplotlib niet beschikbaar of backend-probleem: "
            f"'{MPL_ERR}'. Ik gebruik een fallback (**st.bar_chart**). "
            "Installeer/activeer `matplotlib` voor klassieke plots.")

# Load model (default or user upload)
default_model_path = "Thuiskopie_model.xlsx"
uploaded = st.file_uploader("Upload je Excel-model (optioneel)", type=["xlsx"])

def load_workbook_from_bytes(b):
    bio = io.BytesIO(b.read()) if hasattr(b, "read") else io.BytesIO(b)
    return load_workbook(bio, data_only=True)

if uploaded is not None:
    wb = load_workbook_from_bytes(uploaded)
else:
    wb = load_workbook(default_model_path, data_only=True)

# --- helpers ---
def get_param(ws_params, name, default=None):
    for r in range(2, ws_params.max_row+1):
        if ws_params.cell(row=r, column=1).value == name:
            val = ws_params.cell(row=r, column=3).value
            try:
                return float(val)
            except Exception:
                return val if val is not None else default
    return default

def normalize_dict(dct):
    tot = float(sum(dct.values()))
    return {k: (v/tot if tot>0 else 0.0) for k,v in dct.items()}

def device_factor_of(d, device_pct, dev_w, w_sp_beeld):
    s = device_pct[d]
    sp_mult = w_sp_beeld if d=="Beeld" else 1.0
    return (
        s["Desktop/Laptop"]*dev_w["Desktop/Laptop"] +
        s["Smartphone"]*dev_w["Smartphone"]*sp_mult +
        s["Tablet"]*dev_w["Tablet"] +
        s["E-reader"]*dev_w["E-reader"] +
        s["Externe HDD/NAS"]*dev_w["Externe HDD/NAS"]
    )

def plot_series(series, title):
    if HAVE_MPL:
        fig, ax = plt.subplots()
        ax.bar(series.index, series.values)
        ax.set_title(title)
        ax.set_ylabel("Aandeel")
        ax.set_ylim(0, 1.0)
        st.pyplot(fig)
    else:
        st.bar_chart(series)

# --- read workbook ---
ws_params = wb["Parameters"]
disciplines = ["Audio","AV","Geschriften","Beeld"]
device_cols = ["Desktop/Laptop","Smartphone","Tablet","E-reader","Externe HDD/NAS"]

# Baseline
ws_base = wb["Baseline"]
T = {}
for i, d in enumerate(disciplines):
    v = ws_base.cell(row=5+i, column=2).value  # B5..B8
    T[d] = float(v)/100.0

device_pct = {}
for i, d in enumerate(disciplines):
    r = 14+i
    device_pct[d] = {
        "Desktop/Laptop": float(ws_base.cell(row=r, column=2).value)/100.0,
        "Smartphone":     float(ws_base.cell(row=r, column=3).value)/100.0,
        "Tablet":         float(ws_base.cell(row=r, column=4).value)/100.0,
        "E-reader":       float(ws_base.cell(row=r, column=5).value)/100.0,
        "Externe HDD/NAS":float(ws_base.cell(row=r, column=6).value)/100.0,
    }

# Invoer_2023 shares
ws23 = wb["Invoer_2023"]
share_2023 = {}
vals_2023 = {}
for i, d in enumerate(disciplines):
    v = ws23.cell(row=4+i, column=2).value  # B4..B7
    vals_2023[d] = float(v) if v is not None else None
total_2023 = sum([v for v in vals_2023.values() if v is not None]) or 0.0
for d in disciplines:
    share_2023[d] = (vals_2023[d] / total_2023) if (vals_2023[d] is not None and total_2023>0) else float("nan")

# Parameters
w_trend_A   = get_param(ws_params, "w_trend_A", 0.20)
w_val_A     = get_param(ws_params, "w_valuation_A", 0.20)
w_drag_A    = get_param(ws_params, "w_drager_A", 0.20)
w_fore_A    = get_param(ws_params, "w_foreign_A", 0.20)
w_eq_A      = get_param(ws_params, "w_equal_A", 0.20)

w_trend_B   = get_param(ws_params, "w_trend_B", 0.35)
w_val_B     = get_param(ws_params, "w_valuation_B", 0.35)
w_drag_B    = get_param(ws_params, "w_drager_B", 0.20)
w_fore_B    = get_param(ws_params, "w_foreign_B", 0.10)
bodem_beeld = get_param(ws_params, "bodem_beeld_percent", 10.0) / 100.0

w_trend_C   = get_param(ws_params, "w_trend_C", 0.70)
w_drager_C  = get_param(ws_params, "w_drager_C", 0.30)

dev_w = {
    "Desktop/Laptop": get_param(ws_params, "dev_w_desktop", 0.67),
    "Smartphone":     get_param(ws_params, "dev_w_smartphone", 0.65),
    "Tablet":         get_param(ws_params, "dev_w_tablet", 0.35),
    "E-reader":       get_param(ws_params, "dev_w_ereader", 0.48),
    "Externe HDD/NAS":get_param(ws_params, "dev_w_ext", 1.00),
}
w_sp_beeld = get_param(ws_params, "w_smartphone_beeld", 1.10)

factors_foreign = {
    "Audio": get_param(ws_params, "foreign_factor_audio", 1.0),
    "AV": get_param(ws_params, "foreign_factor_av", 1.0),
    "Geschriften": get_param(ws_params, "foreign_factor_geschriften", 1.0),
    "Beeld": get_param(ws_params, "foreign_factor_beeld", 1.0),
}

ppk   = get_param(ws_params, "ppk_beeld_factor", 1.20)
longv = get_param(ws_params, "longevity_factor_beeld", 1.10)
emb   = get_param(ws_params, "embed_factor_beeld", 1.15)
cld   = get_param(ws_params, "cloud_correction_beeld", 1.05)
msg   = get_param(ws_params, "messenger_cache_factor_beeld", 1.05)
master= get_param(ws_params, "embed_cloud_factor_beeld", 1.00)

target = get_param(ws_params, "beeld_target_percent", 25.0)/100.0
cap = get_param(ws_params, "stability_cap_pp", 0.10)

# ---- Component shares ----
val_factor = {d: 1.0 for d in disciplines}
val_factor["Beeld"] = ppk * longv * emb * cld * msg * master
val_raw = {d: T[d]*val_factor[d] for d in disciplines}
val_norm = normalize_dict(val_raw)

devF = {d: device_factor_of(d, device_pct, dev_w, w_sp_beeld) for d in disciplines}
drager_raw = {d: T[d]*devF[d] for d in disciplines}
drager_norm = normalize_dict(drager_raw)

foreign_raw = {d: T[d]*factors_foreign[d] for d in disciplines}
foreign_norm = normalize_dict(foreign_raw)

equal_share = {d: 0.25 for d in disciplines}

# ---- Scenarios ----
share_A = {d: w_trend_A*T[d] + w_val_A*val_norm[d] + w_drag_A*drager_norm[d] + w_fore_A*foreign_norm[d] + w_eq_A*equal_share[d] for d in disciplines}
share_A = normalize_dict(share_A)

evidence = {d: w_trend_B*T[d] + w_val_B*val_norm[d] + w_drag_B*drager_norm[d] + w_fore_B*foreign_norm[d] for d in disciplines}
evidence = normalize_dict(evidence)
share_B = {}
for d in disciplines:
    if d == "Beeld":
        share_B[d] = bodem_beeld + (1 - bodem_beeld) * evidence[d]
    else:
        share_B[d] = (1 - bodem_beeld) * evidence[d]
share_B = normalize_dict(share_B)

dragerC_raw = {}
for d in disciplines:
    base = share_2023[d] if not math.isnan(share_2023[d]) else T[d]
    dragerC_raw[d] = base * devF[d]
dragerC_norm = normalize_dict(dragerC_raw)
share_C = {d: w_trend_C*(share_2023[d] if not math.isnan(share_2023[d]) else T[d]) + w_drager_C*dragerC_norm[d] for d in disciplines}
share_C = normalize_dict(share_C)

# Stabilisatie (Scenario B vs 2023)
stab_pre = {}
for d in disciplines:
    base = share_2023[d] if not math.isnan(share_2023[d]) else T[d]
    delta = share_B[d] - base
    adj = math.copysign(min(abs(delta), cap), delta)
    stab_pre[d] = base + adj
stab = normalize_dict(stab_pre)

# ---- Output ----
def to_df(dct, name):
    return pd.DataFrame({"Discipline": list(dct.keys()), name: list(dct.values())}).set_index("Discipline")

dfA = to_df(share_A, "Scenario A")
dfB = to_df(share_B, "Scenario B")
dfC = to_df(share_C, "Scenario C")
dfS = to_df(stab, "B gestabiliseerd")
df_all = dfA.join(dfB).join(dfS).join(dfC)

st.subheader("Eindverdeling per scenario")
st.dataframe((df_all*100).round(2))

col1, col2 = st.columns(2)
with col1:
    plot_series(dfB["Scenario B"], "Scenario B")
with col2:
    plot_series(dfS["B gestabiliseerd"], "Scenario B (gestabiliseerd)")

st.markdown("---")
st.subheader("Target helper — Beeld")
st.write(f"Doel (Parameters): **{target*100:.1f}%**")
st.write(f"Huidig Scenario B — Beeld: **{dfB.loc['Beeld','Scenario B']*100:.2f}%**")
st.write(f"Gap: **{(target - dfB.loc['Beeld','Scenario B'])*100:.2f} pp**")

st.markdown("### Download resultaten")
csv = (df_all*100).round(3).to_csv().encode("utf-8")
st.download_button("Download CSV (percentages)", data=csv, file_name="thuiskopie_scenario_resultaten.csv", mime="text/csv")

st.markdown("---")
st.caption("Let op: Dit is een parametrische simulator. Gebruik gevalideerde exports en documenteer parameterkeuzes.")
